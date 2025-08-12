import sqlite3
from datetime import datetime, timedelta
import os
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.popup import Popup
from kivy.uix.spinner import Spinner
from kivy.metrics import dp
from fpdf import FPDF
import openpyxl

# Database setup
def init_db():
    db_path = os.path.join('/sdcard', 'business.db')  # Ruta para Android
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    
    # Users table
    c.execute('''CREATE TABLE IF NOT EXISTS users
                 (username TEXT PRIMARY KEY, password TEXT)''')
    
    # Sales table
    c.execute('''CREATE TABLE IF NOT EXISTS sales
                 (id INTEGER PRIMARY KEY, date TEXT, value REAL, payment_method TEXT)''')
    
    # Purchases table
    c.execute('''CREATE TABLE IF NOT EXISTS purchases
                 (id INTEGER PRIMARY KEY, date TEXT, place TEXT, product TEXT, quantity REAL, value REAL)''')
    
    # Places history
    c.execute('''CREATE TABLE IF NOT EXISTS places
                 (name TEXT UNIQUE)''')
    
    # Products history
    c.execute('''CREATE TABLE IF NOT EXISTS products
                 (name TEXT UNIQUE)''')
    
    conn.commit()
    conn.close()
    return db_path

# Login Screen
class LoginScreen(BoxLayout):
    def __init__(self, app, **kwargs):
        super().__init__(**kwargs)
        self.app = app
        self.orientation = 'vertical'
        self.padding = dp(20)
        self.spacing = dp(10)
        self.db_path = init_db()
        
        # Username input
        self.add_widget(Label(text="Usuario:", font_size=dp(20), color=(0, 0, 0, 1)))
        self.user_input = TextInput(font_size=dp(20), multiline=False)
        self.add_widget(self.user_input)
        
        # Password input
        self.add_widget(Label(text="Contraseña:", font_size=dp(20), color=(0, 0, 0, 1)))
        self.pass_input = TextInput(font_size=dp(20), multiline=False, password=True)
        self.add_widget(self.pass_input)
        
        # Buttons
        self.login_button = Button(text="Iniciar Sesión", size_hint=(1, 0.2), background_color=(1, 0.84, 0, 1))
        self.login_button.bind(on_press=self.login)
        self.add_widget(self.login_button)
        
        self.register_button = Button(text="Registrar", size_hint=(1, 0.2), background_color=(1, 0.84, 0, 1))
        self.register_button.bind(on_press=self.register)
        self.add_widget(self.register_button)

    def login(self, instance):
        username = self.user_input.text
        password = self.pass_input.text
        
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        c.execute("SELECT * FROM users WHERE username=? AND password=?", (username, password))
        user = c.fetchone()
        conn.close()
        
        if user:
            self.app.username = username
            self.app.root.clear_widgets()
            self.app.root.add_widget(MainMenuScreen(self.app))
        else:
            self.show_popup("Error", "Credenciales inválidas")

    def register(self, instance):
        username = self.user_input.text
        password = self.pass_input.text
        
        if not username or not password:
            self.show_popup("Error", "Campos requeridos")
            return
        
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        try:
            c.execute("INSERT INTO users (username, password) VALUES (?, ?)", (username, password))
            conn.commit()
            self.show_popup("Éxito", "Usuario registrado")
        except sqlite3.IntegrityError:
            self.show_popup("Error", "Usuario ya existe")
        conn.close()

    def show_popup(self, title, message):
        popup = Popup(title=title, content=Label(text=message), size_hint=(0.8, 0.4))
        popup.open()

# Main Menu Screen
class MainMenuScreen(BoxLayout):
    def __init__(self, app, **kwargs):
        super().__init__(**kwargs)
        self.app = app
        self.orientation = 'vertical'
        self.padding = dp(20)
        self.spacing = dp(10)
        
        self.add_widget(Label(text=f"Bienvenido, {self.app.username}", font_size=dp(24), color=(0, 0, 0, 1)))
        
        sales_button = Button(text="Ventas", size_hint=(1, 0.2), background_color=(1, 0.84, 0, 1))
        sales_button.bind(on_press=self.go_to_sales)
        self.add_widget(sales_button)
        
        purchases_button = Button(text="Compras", size_hint=(1, 0.2), background_color=(1, 0.84, 0, 1))
        purchases_button.bind(on_press=self.go_to_purchases)
        self.add_widget(purchases_button)
        
        balance_button = Button(text="Balance e Informes", size_hint=(1, 0.2), background_color=(1, 0.84, 0, 1))
        balance_button.bind(on_press=self.go_to_balance)
        self.add_widget(balance_button)
        
        logout_button = Button(text="Cerrar Sesión", size_hint=(1, 0.2), background_color=(1, 0.84, 0, 1))
        logout_button.bind(on_press=self.logout)
        self.add_widget(logout_button)

    def go_to_sales(self, instance):
        self.app.root.clear_widgets()
        self.app.root.add_widget(SalesScreen(self.app))

    def go_to_purchases(self, instance):
        self.app.root.clear_widgets()
        self.app.root.add_widget(PurchasesScreen(self.app))

    def go_to_balance(self, instance):
        self.app.root.clear_widgets()
        self.app.root.add_widget(BalanceScreen(self.app))

    def logout(self, instance):
        self.app.root.clear_widgets()
        self.app.root.add_widget(LoginScreen(self.app))

# Sales Screen
class SalesScreen(BoxLayout):
    def __init__(self, app, **kwargs):
        super().__init__(**kwargs)
        self.app = app
        self.orientation = 'vertical'
        self.padding = dp(20)
        self.spacing = dp(10)
        self.db_path = app.db_path
        
        today = datetime.now().strftime("%Y-%m-%d")
        self.add_widget(Label(text=f"Fecha: {today}", font_size=dp(20), color=(0, 0, 0, 1)))
        
        self.add_widget(Label(text="Valor de la Venta:", font_size=dp(16), color=(0, 0, 0, 1)))
        self.sale_value = TextInput(font_size=dp(16), multiline=False)
        self.add_widget(self.sale_value)
        
        self.add_widget(Label(text="Modo de Pago:", font_size=dp(16), color=(0, 0, 0, 1)))
        self.payment_method = Spinner(
            text="Efectivo",
            values=["Efectivo", "Nequi", "Transferencia", "Deuda"],
            font_size=dp(16)
        )
        self.add_widget(self.payment_method)
        
        add_button = Button(text="Agregar Venta", size_hint=(1, 0.2), background_color=(1, 0.84, 0, 1))
        add_button.bind(on_press=self.add_sale)
        self.add_widget(add_button)
        
        self.sales_list = Label(text="", font_size=dp(14), color=(0, 0, 0, 1), text_size=(None, None), size_hint=(1, 1))
        self.add_widget(self.sales_list)
        
        back_button = Button(text="Volver", size_hint=(1, 0.2), background_color=(1, 0.84, 0, 1))
        back_button.bind(on_press=self.back_to_menu)
        self.add_widget(back_button)
        
        self.load_daily_sales()

    def add_sale(self, instance):
        value = self.sale_value.text
        method = self.payment_method.text
        if not value or not method:
            self.show_popup("Error", "Campos requeridos")
            return
        try:
            value = float(value)
        except ValueError:
            self.show_popup("Error", "Valor inválido")
            return
        
        today = datetime.now().strftime("%Y-%m-%d")
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        c.execute("INSERT INTO sales (date, value, payment_method) VALUES (?, ?, ?)", (today, value, method))
        conn.commit()
        conn.close()
        
        self.show_popup("Éxito", "Venta agregada")
        self.load_daily_sales()

    def load_daily_sales(self):
        self.sales_list.text = ""
        today = datetime.now().strftime("%Y-%m-%d")
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        c.execute("SELECT payment_method, SUM(value) FROM sales WHERE date=? GROUP BY payment_method", (today,))
        results = c.fetchall()
        total = 0
        text = ""
        for method, val in results:
            text += f"{method}: {val}\n"
            total += val
        text += f"Total: {total}"
        self.sales_list.text = text
        conn.close()

    def show_popup(self, title, message):
        popup = Popup(title=title, content=Label(text=message), size_hint=(0.8, 0.4))
        popup.open()

    def back_to_menu(self, instance):
        self.app.root.clear_widgets()
        self.app.root.add_widget(MainMenuScreen(self.app))

# Purchases Screen
class PurchasesScreen(BoxLayout):
    def __init__(self, app, **kwargs):
        super().__init__(**kwargs)
        self.app = app
        self.orientation = 'vertical'
        self.padding = dp(20)
        self.spacing = dp(10)
        self.db_path = app.db_path
        
        today = datetime.now().strftime("%Y-%m-%d")
        self.add_widget(Label(text=f"Fecha: {today}", font_size=dp(20), color=(0, 0, 0, 1)))
        
        self.add_widget(Label(text="Lugar de Compra:", font_size=dp(16), color=(0, 0, 0, 1)))
        self.place_input = TextInput(font_size=dp(16), multiline=False)
        self.add_widget(self.place_input)
        
        self.add_widget(Label(text="Producto:", font_size=dp(16), color=(0, 0, 0, 1)))
        self.product_input = TextInput(font_size=dp(16), multiline=False)
        self.add_widget(self.product_input)
        
        self.add_widget(Label(text="Cantidad/Peso:", font_size=dp(16), color=(0, 0, 0, 1)))
        self.quantity = TextInput(font_size=dp(16), multiline=False)
        self.add_widget(self.quantity)
        
        self.add_widget(Label(text="Valor:", font_size=dp(16), color=(0, 0, 0, 1)))
        self.purchase_value = TextInput(font_size=dp(16), multiline=False)
        self.add_widget(self.purchase_value)
        
        add_button = Button(text="Agregar Producto", size_hint=(1, 0.2), background_color=(1, 0.84, 0, 1))
        add_button.bind(on_press=self.add_purchase)
        self.add_widget(add_button)
        
        self.purchases_list = Label(text="", font_size=dp(14), color=(0, 0, 0, 1), text_size=(None, None), size_hint=(1, 1))
        self.add_widget(self.purchases_list)
        
        back_button = Button(text="Volver", size_hint=(1, 0.2), background_color=(1, 0.84, 0, 1))
        back_button.bind(on_press=self.back_to_menu)
        self.add_widget(back_button)
        
        self.load_daily_purchases()

    def add_purchase(self, instance):
        place = self.place_input.text
        product = self.product_input.text
        qty = self.quantity.text
        value = self.purchase_value.text
        
        if not all([place, product, qty, value]):
            self.show_popup("Error", "Campos requeridos")
            return
        
        try:
            qty = float(qty)
            value = float(value)
        except ValueError:
            self.show_popup("Error", "Valores inválidos")
            return
        
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        
        # Add place if new
        try:
            c.execute("INSERT INTO places (name) VALUES (?)", (place,))
            conn.commit()
        except sqlite3.IntegrityError:
            pass
        
        # Add product if new
        try:
            c.execute("INSERT INTO products (name) VALUES (?)", (product,))
            conn.commit()
        except sqlite3.IntegrityError:
            pass
        
        today = datetime.now().strftime("%Y-%m-%d")
        c.execute("INSERT INTO purchases (date, place, product, quantity, value) VALUES (?, ?, ?, ?, ?)",
                  (today, place, product, qty, value))
        conn.commit()
        conn.close()
        
        self.show_popup("Éxito", "Compra agregada")
        self.load_daily_purchases()

    def load_daily_purchases(self):
        self.purchases_list.text = ""
        today = datetime.now().strftime("%Y-%m-%d")
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        c.execute("SELECT place, product, quantity, value FROM purchases WHERE date=?", (today,))
        results = c.fetchall()
        total = 0
        text = ""
        for row in results:
            text += f"{row[0]} - {row[1]}: {row[2]} @ {row[3]}\n"
            total += row[3]
        text += f"Total: {total}"
        self.purchases_list.text = text
        conn.close()

    def show_popup(self, title, message):
        popup = Popup(title=title, content=Label(text=message), size_hint=(0.8, 0.4))
        popup.open()

    def back_to_menu(self, instance):
        self.app.root.clear_widgets()
        self.app.root.add_widget(MainMenuScreen(self.app))

# Balance Screen
class BalanceScreen(BoxLayout):
    def __init__(self, app, **kwargs):
        super().__init__(**kwargs)
        self.app = app
        self.orientation = 'vertical'
        self.padding = dp(20)
        self.spacing = dp(10)
        self.db_path = app.db_path
        
        self.add_widget(Label(text="Balance e Informes", font_size=dp(24), color=(0, 0, 0, 1)))
        
        self.add_widget(Label(text="Tipo de Informe:", font_size=dp(16), color=(0, 0, 0, 1)))
        self.report_type = Spinner(
            text="Diario",
            values=["Diario", "Semanal", "Quincenal", "Mensual", "Anual"],
            font_size=dp(16)
        )
        self.add_widget(self.report_type)
        
        self.add_widget(Label(text="Fecha Inicio (YYYY-MM-DD):", font_size=dp(16), color=(0, 0, 0, 1)))
        self.start_date = TextInput(font_size=dp(16), multiline=False)
        self.add_widget(self.start_date)
        
        self.add_widget(Label(text="Fecha Fin (YYYY-MM-DD):", font_size=dp(16), color=(0, 0, 0, 1)))
        self.end_date = TextInput(font_size=dp(16), multiline=False)
        self.add_widget(self.end_date)
        
        generate_button = Button(text="Generar Informe", size_hint=(1, 0.2), background_color=(1, 0.84, 0, 1))
        generate_button.bind(on_press=self.generate_report)
        self.add_widget(generate_button)
        
        self.report_text = Label(text="", font_size=dp(14), color=(0, 0, 0, 1), text_size=(None, None), size_hint=(1, 1))
        self.add_widget(self.report_text)
        
        pdf_button = Button(text="Exportar a PDF", size_hint=(1, 0.2), background_color=(1, 0.84, 0, 1))
        pdf_button.bind(on_press=self.export_pdf)
        self.add_widget(pdf_button)
        
        excel_button = Button(text="Exportar a Excel", size_hint=(1, 0.2), background_color=(1, 0.84, 0, 1))
        excel_button.bind(on_press=self.export_excel)
        self.add_widget(excel_button)
        
        back_button = Button(text="Volver", size_hint=(1, 0.2), background_color=(1, 0.84, 0, 1))
        back_button.bind(on_press=self.back_to_menu)
        self.add_widget(back_button)

    def generate_report(self, instance):
        report_type = self.report_type.text
        start = self.start_date.text
        end = self.end_date.text
        
        if not report_type:
            self.show_popup("Error", "Seleccione tipo de informe")
            return
        
        today = datetime.now()
        if report_type == "Diario":
            start = end = today.strftime("%Y-%m-%d")
        elif report_type == "Semanal":
            start = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
            end = (today + timedelta(days=6 - today.weekday())).strftime("%Y-%m-%d")
        elif report_type == "Quincenal":
            if today.day <= 15:
                start = today.replace(day=1).strftime("%Y-%m-%d")
                end = today.replace(day=15).strftime("%Y-%m-%d")
            else:
                start = today.replace(day=16).strftime("%Y-%m-%d")
                end = (today.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
                end = end.strftime("%Y-%m-%d")
        elif report_type == "Mensual":
            start = today.replace(day=1).strftime("%Y-%m-%d")
            end = (today.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            end = end.strftime("%Y-%m-%d")
        elif report_type == "Anual":
            start = today.replace(month=1, day=1).strftime("%Y-%m-%d")
            end = today.replace(month=12, day=31).strftime("%Y-%m-%d")
        
        if start == "" or end == "":
            self.show_popup("Error", "Ingrese fechas válidas")
            return
        
        try:
            datetime.strptime(start, "%Y-%m-%d")
            datetime.strptime(end, "%Y-%m-%d")
        except ValueError:
            self.show_popup("Error", "Formato de fecha inválido")
            return
        
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        
        # Sales
        c.execute("SELECT payment_method, SUM(value) FROM sales WHERE date BETWEEN ? AND ? GROUP BY payment_method", (start, end))
        sales = c.fetchall()
        total_sales = sum([row[1] for row in sales])
        
        # Purchases
        c.execute("SELECT SUM(value) FROM purchases WHERE date BETWEEN ? AND ?", (start, end))
        total_purchases = c.fetchone()[0] or 0
        
        profits = total_sales - total_purchases
        
        text = f"Informe {report_type} ({start} a {end})\n\n"
        text += "Ventas:\n"
        for method, val in sales:
            text += f"{method}: {val}\n"
        text += f"Total Ventas: {total_sales}\n\n"
        text += f"Compras Total: {total_purchases}\n\n"
        text += f"Ganancias: {profits}\n"
        self.report_text.text = text
        
        self.current_report = {
            'type': report_type,
            'start': start,
            'end': end,
            'sales': sales,
            'total_sales': total_sales,
            'total_purchases': total_purchases,
            'profits': profits
        }
        
        conn.close()

    def export_pdf(self, instance):
        if not hasattr(self, 'current_report'):
            self.show_popup("Error", "Genere un informe primero")
            return
        
        report = self.current_report
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt=f"Informe {report['type']} ({report['start']} a {report['end']})", ln=1, align='C')
        pdf.cell(200, 10, txt="Ventas:", ln=1)
        for method, val in report['sales']:
            pdf.cell(200, 10, txt=f"{method}: {val}", ln=1)
        pdf.cell(200, 10, txt=f"Total Ventas: {report['total_sales']}", ln=1)
        pdf.cell(200, 10, txt=f"Compras Total: {report['total_purchases']}", ln=1)
        pdf.cell(200, 10, txt=f"Ganancias: {report['profits']}", ln=1)
        
        file_path = os.path.join('/sdcard/Documents', f"report_{report['type']}.pdf")
        pdf.output(file_path)
        self.show_popup("Éxito", f"Exportado a {file_path}")

    def export_excel(self, instance):
        if not hasattr(self, 'current_report'):
            self.show_popup("Error", "Genere un informe primero")
            return
        
        report = self.current_report
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Informe"
        
        ws['A1'] = f"Informe {report['type']} ({report['start']} a {report['end']})"
        ws['A3'] = "Ventas"
        row = 4
        for method, val in report['sales']:
            ws[f'A{row}'] = method
            ws[f'B{row}'] = val
            row += 1
        ws[f'A{row}'] = "Total Ventas"
        ws[f'B{row}'] = report['total_sales']
        row += 2
        ws[f'A{row}'] = "Compras Total"
        ws[f'B{row}'] = report['total_purchases']
        row += 1
        ws[f'A{row}'] = "Ganancias"
        ws[f'B{row}'] = report['profits']
        
        file_path = os.path.join('/sdcard/Documents', f"report_{report['type']}.xlsx")
        wb.save(file_path)
        self.show_popup("Éxito", f"Exportado a {file_path}")

    def show_popup(self, title, message):
        popup = Popup(title=title, content=Label(text=message), size_hint=(0.8, 0.4))
        popup.open()

    def back_to_menu(self, instance):
        self.app.root.clear_widgets()
        self.app.root.add_widget(MainMenuScreen(self.app))

# Main App
class BusinessApp(App):
    def __init__(self):
        super().__init__()
        self.username = None
        self.db_path = init_db()
    
    def build(self):
        return LoginScreen(self)

if __name__ == '__main__':
    BusinessApp().run()